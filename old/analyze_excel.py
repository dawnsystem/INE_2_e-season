import pandas as pd
import os
from openpyxl import load_workbook

def analyze_excel_structure():
    folder_path = r"C:\Users\david\Downloads\ine_transformer"
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    
    for file in excel_files:
        file_path = os.path.join(folder_path, file)
        print(f"\n{'='*80}")
        print(f"ARCHIVO: {file}")
        print('='*80)
        
        # Cargar el workbook para ver las hojas
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        print(f"Hojas disponibles: {sheet_names}")
        wb.close()
        
        # Analizar cada hoja
        for sheet in sheet_names:
            print(f"\n--- Hoja: {sheet} ---")
            try:
                df = pd.read_excel(file_path, sheet_name=sheet)
                print(f"Dimensiones: {df.shape[0]} filas x {df.shape[1]} columnas")
                print(f"Columnas: {list(df.columns)}")
                
                # Mostrar primeras filas para entender estructura
                print("\nPrimeras 5 filas:")
                print(df.head())
                
                # Información sobre tipos de datos
                print("\nTipos de datos:")
                print(df.dtypes)
                
                # Valores únicos en columnas clave
                for col in df.columns:
                    if df[col].dtype == 'object':  # Solo para columnas de texto
                        unique_count = df[col].nunique()
                        if unique_count < 20:  # Solo si hay pocos valores únicos
                            print(f"\nValores únicos en '{col}': {df[col].unique()}")
                
            except Exception as e:
                print(f"Error al leer hoja {sheet}: {e}")

if __name__ == "__main__":
    analyze_excel_structure()